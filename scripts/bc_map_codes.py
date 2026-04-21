"""
Lees een BC stock-export (Excel) in, zoek voor elke oude BC-code de nieuwe
variant op via de BC item mapping, en schrijf het resultaat weg als een
pipe-separated filter string (FP001|FP002|FP003...).

Deze output kan direct in Business Central gebruikt worden als filter.

Gebruik:
    python scripts/bc_map_codes.py <excel_path>
    python scripts/bc_map_codes.py <excel_path> --column "No."
    python scripts/bc_map_codes.py <excel_path> --mapping "C:/pad/naar/mapping.xlsx"
    python scripts/bc_map_codes.py <excel_path> --api https://prodwilrijk.be

Standaard:
    - Leest kolom "No." uit de eerste sheet.
    - Haalt de mapping via https://prodwilrijk.be/api/bc-mappings.
    - Schrijft de output naar <excel_basename>_bc36_filter.txt naast het
      input-bestand en print hem ook in het scherm.

Met --mapping kun je een lokaal mapping-Excel bestand gebruiken in plaats van
de API; het verwacht kolommen "Oud" en "Nieuw" (of "old_code" / "new_code").
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import urllib.request
from typing import Dict, Iterable, List, Tuple

try:
    import openpyxl
except ImportError:
    print("openpyxl is niet geïnstalleerd. Installeer met: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


DEFAULT_API_BASE = "https://prodwilrijk.be"


def fetch_mapping_from_api(base_url: str) -> Dict[str, str]:
    """Haal de volledige mapping-tabel op en bouw oud→nieuw dict (uppercase keys)."""
    url = base_url.rstrip("/") + "/api/bc-mappings"
    print(f"[info] Mapping ophalen van {url} ...", file=sys.stderr)
    req = urllib.request.Request(url, headers={"Accept": "application/json"})
    with urllib.request.urlopen(req, timeout=60) as resp:
        payload = json.loads(resp.read().decode("utf-8"))
    rows = payload.get("mappings") or []
    mapping: Dict[str, str] = {}
    for r in rows:
        old = (r.get("old_code") or "").strip()
        new = (r.get("new_code") or "").strip()
        if old and new:
            mapping[old.upper()] = new
    print(f"[info] {len(mapping)} mappings geladen.", file=sys.stderr)
    return mapping


def fetch_mapping_from_excel(path: str) -> Dict[str, str]:
    """Leest mapping uit een lokaal Excel bestand.

    Accepteert kolomnamen "Oud"/"Nieuw", "oud_code"/"nieuw_code",
    "old_code"/"new_code" (case-insensitive). Pakt de eerste match.
    """
    print(f"[info] Mapping lezen van {path} ...", file=sys.stderr)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip().lower() if h is not None else "" for h in next(rows)]

    def find_col(candidates: Iterable[str]) -> int:
        for idx, h in enumerate(headers):
            if h in candidates:
                return idx
        raise KeyError(f"Geen kolom gevonden voor {candidates!r} in mapping file.")

    old_idx = find_col({"oud", "old", "old_code", "oud_code", "oude bc code", "oude code"})
    new_idx = find_col({"nieuw", "new", "new_code", "nieuw_code", "nieuwe bc code", "nieuwe code"})

    mapping: Dict[str, str] = {}
    for row in rows:
        if row is None:
            continue
        old = row[old_idx]
        new = row[new_idx]
        if old is None or new is None:
            continue
        old_s = str(old).strip()
        new_s = str(new).strip()
        if old_s and new_s:
            mapping[old_s.upper()] = new_s
    print(f"[info] {len(mapping)} mappings geladen.", file=sys.stderr)
    return mapping


def read_codes(excel_path: str, column: str, sheet: str | None = None) -> List[str]:
    """Leest alle niet-lege codes uit de opgegeven kolom."""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []

    target = column.strip().lower()
    try:
        col_idx = next(i for i, h in enumerate(headers) if h.lower() == target)
    except StopIteration:
        print(f"[fout] Kolom '{column}' niet gevonden. Beschikbaar: {headers}", file=sys.stderr)
        sys.exit(1)

    codes: List[str] = []
    for row in rows:
        if row is None:
            continue
        v = row[col_idx]
        if v is None:
            continue
        s = str(v).strip()
        if s:
            codes.append(s)
    return codes


def translate(
    codes: List[str], mapping: Dict[str, str]
) -> Tuple[List[str], List[str], List[str]]:
    """Returnt (unieke nieuwe codes in volgorde, dubbele inputs, niet-gemapte inputs)."""
    new_codes: List[str] = []
    seen: set[str] = set()
    unmapped: List[str] = []
    duplicates: List[str] = []
    seen_raw: set[str] = set()

    for code in codes:
        key = code.upper()
        if key in seen_raw:
            duplicates.append(code)
            continue
        seen_raw.add(key)

        new = mapping.get(key)
        if not new:
            unmapped.append(code)
            continue
        if new.upper() in seen:
            continue
        seen.add(new.upper())
        new_codes.append(new)
    return new_codes, duplicates, unmapped


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("excel", help="Pad naar de BC stock-export (.xlsx)")
    parser.add_argument("--column", default="No.", help='Kolomnaam met de oude codes (default: "No.")')
    parser.add_argument("--sheet", default=None, help="Naam van de sheet (default: eerste sheet)")
    parser.add_argument("--mapping", default=None, help="Optioneel lokaal mapping Excel bestand")
    parser.add_argument("--api", default=DEFAULT_API_BASE, help="Base URL voor API (default: %(default)s)")
    parser.add_argument("--output", default=None, help="Pad voor output .txt (default: naast input bestand)")
    parser.add_argument("--separator", default="|", help='Separator voor de output (default: "|")')
    parser.add_argument("--keep-unmapped", action="store_true", help="Onbekende codes tóch opnemen (ongewijzigd)")
    args = parser.parse_args()

    if not os.path.isfile(args.excel):
        print(f"[fout] Bestand niet gevonden: {args.excel}", file=sys.stderr)
        return 1

    if args.mapping:
        mapping = fetch_mapping_from_excel(args.mapping)
    else:
        try:
            mapping = fetch_mapping_from_api(args.api)
        except Exception as exc:
            print(f"[fout] Kon mapping niet ophalen via API: {exc}", file=sys.stderr)
            print("       Gebruik --mapping <excel> voor een lokaal bestand.", file=sys.stderr)
            return 1

    if not mapping:
        print("[fout] Mapping is leeg — niets te doen.", file=sys.stderr)
        return 1

    codes = read_codes(args.excel, column=args.column, sheet=args.sheet)
    if not codes:
        print("[fout] Geen codes gevonden in het bestand.", file=sys.stderr)
        return 1
    print(f"[info] {len(codes)} rijen gelezen uit {args.excel} (kolom '{args.column}').", file=sys.stderr)

    new_codes, duplicates, unmapped = translate(codes, mapping)
    if args.keep_unmapped:
        # Voeg onbekende codes ongewijzigd toe achteraan, dedup.
        seen = {c.upper() for c in new_codes}
        for u in unmapped:
            if u.upper() not in seen:
                new_codes.append(u)
                seen.add(u.upper())

    filter_string = args.separator.join(new_codes)

    out_path = args.output or os.path.splitext(args.excel)[0] + "_bc36_filter.txt"
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(filter_string)

    # Rapport naar stderr zodat stdout puur de filter-string blijft.
    print("", file=sys.stderr)
    print("========= RESULTAAT =========", file=sys.stderr)
    print(f"  Unieke oude codes ingelezen : {len(codes) - len(duplicates)}", file=sys.stderr)
    print(f"  Dubbelen in input           : {len(duplicates)}", file=sys.stderr)
    print(f"  Gemapt → nieuw              : {len(new_codes) - (len(unmapped) if args.keep_unmapped else 0)}", file=sys.stderr)
    print(f"  Niet gemapt                 : {len(unmapped)}", file=sys.stderr)
    if unmapped:
        preview = ", ".join(unmapped[:15])
        suffix = " ..." if len(unmapped) > 15 else ""
        print(f"    → voorbeeld: {preview}{suffix}", file=sys.stderr)
    print(f"  Output geschreven naar      : {out_path}", file=sys.stderr)
    print(f"  Totaal in filter            : {len(new_codes)} codes, {len(filter_string)} tekens", file=sys.stderr)
    print("=============================", file=sys.stderr)
    print("", file=sys.stderr)

    # Feitelijke filter-string naar stdout (pipen, kopiëren, ...).
    print(filter_string)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
